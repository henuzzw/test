import pickle
import os
import sys

import openpyxl as op

K_size=1


def write(epoch,datalist):
    bg = op.load_workbook(r"Grace.xlsx")  # 应先将excel文件放入到工作目录下
    sheet = bg["Sheet"+str(epoch)]
    row=sheet.max_row+1
    #max_row和max_column两个方法
    for col in range(1,len(datalist)+1):
        sheet.cell(row, col,datalist[col-1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save("Grace.xlsx")  # 对文件进行保存

versionNum = {'Lang': 65, 'Time': 27, 'Chart': 26, 'Math': 106, 'Closure': 133, 'Mockito': 38,
                   'CommonsCli': 24, 'CommonsCodec': 22, 'CommonsCsv': 12, 'CommonsJXPath': 14,
                   'JacksonCore': 13, 'JacksonDatabind': 39, 'JacksonXml': 5, 'Jsoup': 63,'tcas':30,'Cli':38,'JxPath':22}
proj = sys.argv[1]
seed = int(sys.argv[2])
lr = float(sys.argv[3])
batch_size = int(sys.argv[4])
modelName = sys.argv[5]
EPOCHS=15
missk=0
data = pickle.load(open(proj + '.pkl', 'rb'))

t = {}
t[0] = []  # 最好排名
t[1] = {}  # 最好排名对应的列表
t[2] = []  # top-1出现的epoch
t[3] = {}  # 所有epoch对应的所有列表

for i in range(0, 10):#len(data)):
    #print("i",i)
    if not os.path.exists(proj + 'res%d_%d_%s_%s.pkl' % ( i, seed, lr, batch_size)):
        print(proj + 'res%d_%d_%s_%s.pkl' % ( i, seed, lr, batch_size))
        continue
    p = pickle.load(open(proj + 'res%d_%d_%s_%s.pkl' % ( i, seed, lr, batch_size), 'rb'))
    p = p[i]
    Max_expoch = [len(p[0]), len(p[1]), len(p[2]), len(p[3])]
    Max_expoch = [1, 1, len(p[2]), EPOCHS]
    for j in [0, 2]:
        for x in range(Max_expoch[j]):
            t[j].append(p[j][x])
#    for key in p[1]:
#        t[1][key] = p[1][key]
    for j in range(Max_expoch[3]):
        #print("j",j,)
        if j not in t[3]:
            t[3][j] = {}
            #############t[1][j]={}
        #print("*******",p[3],p[3][j])
        for key in p[3][j]:
            print("i,j,key",i,j,key)
            t[3][j][key] = p[3][j][key]
            #############t[1][j][key] = p[1][j][key]

open(proj + 'res%d_%s_%s.pkl' % ( seed, lr, batch_size), 'wb').write(pickle.dumps(t))

ANS = {}  # 每一个epoch对应的top12345
MAPs = {}  # MAR
ANS_version = {}
MFRs = {}
anslist = {}
print(t[3])
for epoch in t[3]:
    anslist[epoch] = []
    MAPs[epoch] = 0
    MFRs[epoch] = 0
    ANS_version[epoch] = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    ANS[epoch] = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 1-3-5-10
    ranks = t[3][epoch]
    # num=0
    for key in ranks:
        print(epoch,key)
        minl = 1e9
        # num+=1
        AP = 0
        ranklist = ranks[key]
        #for x in data[key]['ans']:#方法级别的故障定位
        #print(key,dict[proj][key],data[key]['ans'])
        for x in data[key]['ans']:#语句级别的故障定
            rank_x = ranklist.index(x-1)
            minl = min(rank_x, minl)
            AP += (rank_x + 1)
            if rank_x < 10:
                ANS[epoch][rank_x] += 1
        anslist[epoch].append(AP)
        #MAPs[epoch] += (AP / len(data[key]['ans']))#方法级别的故障定位
        MAPs[epoch] += (AP / len(data[key]['ans']))#语句级别的故障定位
        if minl < 10:
            ANS_version[epoch][minl] += 1
        MFRs[epoch] += (minl + 1)
    MAPs[epoch] = (MAPs[epoch] / (len(data)-missk))
    MFRs[epoch] = (MFRs[epoch] /  (len(data)-missk))
print(anslist)

TOP1 = 0
TOP3 = 0
TOP5 = 0
TOP10 = 0
TOP1_v = 0
TOP3_v = 0
TOP5_v = 0
TOP10_v = 0
epochs = len(ANS) - 1
print(epochs)
MAP = 0
MFR = 0
for epoch in ANS:
    top1 = ANS[epoch][0]
    top3 = top1 + ANS[epoch][1] + ANS[epoch][2]
    top5 = top3 + ANS[epoch][3] + ANS[epoch][4]
    top10 = top5 + ANS[epoch][5] + ANS[epoch][6]+ANS[epoch][7] + ANS[epoch][8]+ ANS[epoch][9]
    top1_v = ANS_version[epoch][0]
    top3_v = top1_v + ANS_version[epoch][1] + ANS_version[epoch][2]
    top5_v = top3_v + ANS_version[epoch][3] + ANS_version[epoch][4]
    top10_v = top5_v + ANS_version[epoch][5] + ANS_version[epoch][6]+ANS_version[epoch][7] + ANS_version[epoch][8]+ ANS_version[epoch][9]
    print(epoch, "faults: ", top1, top3, top5, top10, "faulty versions: ", top1_v, top3_v, top5_v, top10_v, "MAP:", MAPs[epoch], "MFR:",
          MFRs[epoch])
    write(epoch, [proj, modelName, top1, top3, top5, top10, top1_v, top3_v, top5_v, top10_v, MAPs[epoch], MFRs[epoch], K_size])
    if epoch == 0:
        continue
    TOP1 += top1
    TOP3 += top3
    TOP5 += top5
    TOP10 += top10
    TOP1_v += top1_v
    TOP3_v += top3_v
    TOP5_v += top5_v
    TOP10_v += top10_v
    MAP += MAPs[epoch]
    MFR += MFRs[epoch]

print(epochs, "avgTOP-faults:", TOP1 / epochs, TOP3 / epochs, TOP5 / epochs, TOP10 / epochs,"average faulty versions: ",
      TOP1_v / epochs, TOP3_v / epochs, TOP5_v / epochs, TOP10_v / epochs, "average MAP", MAP / epochs, "average MFR", MFR / epochs)

tmp = [proj, modelName, TOP1 / epochs, TOP3 / epochs, TOP5 / epochs, TOP10 / epochs, TOP1_v / epochs, TOP3_v / epochs, TOP5_v / epochs, TOP10_v / epochs,
       MAP / epochs, MFR / epochs]
# bestTOP1 = 0
# bestTOP3 = 0
# bestTOP5 = 0
# bestTOP10 = 0
# MFR = 0
# print("len(t[0])", len(t[0]))
# for i in range(len(t[0])):
#     x = t[0][i]
#     if x == 0:
#         bestTOP1 += 1
#     if x < 3:
#         bestTOP3 += 1
#     if x < 5:
#         bestTOP5 += 1
#     if x < 10:
#         bestTOP10 += 1
#     MFR += (x + 1)
# MFR = MFR / (len(data)-missk)
# print("bestTOP-version:", bestTOP1, bestTOP3, bestTOP5, bestTOP10, "best-MFR", MFR)

# ans = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # 对应所有最好的版本的top12345
# num = 0
# MAP = 0
# for key in t[1]:
#     num += 1
#     ranklist = t[1][key]
#     AP = 0
#     #for x in data[key]['ans']:#方法级别的故障定位
#     for x in data[key]['ans']:#语句级别的故障定位
#         rank_x = ranklist.index(x-1)
#         AP += (rank_x + 1)
#         if rank_x < 10:
#             ans[rank_x] += 1
#     #MAP += AP / len(data[key]['ans'])#方法级别的故障定位
#     MAP += AP / len(data[key]['ans'])#语句级别的故障定位
# MAP = MAP /  (len(data)-missk)
# bestfault_TOP1 = ans[0]
# bestfault_TOP3 = bestfault_TOP1 + ans[1] + ans[2]
# bestfault_TOP5 = bestfault_TOP3 + ans[3] + ans[4]
# bestfault_TOP10 = bestfault_TOP5 + ans[5] + ans[6] + ans[7] + ans[8] + ans[9]
# print("bestTOP-faluts:", bestfault_TOP1, bestfault_TOP3, bestfault_TOP5, bestfault_TOP10, "best-MAP", MAP)
# tmp.extend([bestTOP1, bestTOP3, bestTOP5, bestTOP10, MFR, bestfault_TOP1, bestfault_TOP3, bestfault_TOP5, bestfault_TOP10, MAP, K_size])
write(epochs + 1, tmp)













