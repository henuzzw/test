import torch
from torch import optim
from Dataset import SumDataset
import os
from tqdm import tqdm
from Model import *
import numpy as np
#from annoy import AnnoyIndex
from nltk import word_tokenize
import pickle
from ScheduledOptim import ScheduledOptim
from nltk.translate.bleu_score import corpus_bleu
import pandas as pd
import random
import sys
import openpyxl
from datetime import datetime
#import wandb
#wandb.init(project="codesum")
class dotdict(dict):
    def __getattr__(self, name):
        return self[name]

NlLen_map = {"Time":4485,"Math":4569, "Math1":4500, "Lang":510,"Lang1":280, "tcas":160, "Chart": 2318, "Mockito":1780, "unknown":2200,"Cli":583,"JxPath":880}
CodeLen_map={"Time":2239,"Math":2441,"Math1":2700, "Lang":291,"Lang1":300, "tcas":220, "Chart": 4821, "Mockito":1176, "unknown":2800,"Cli":522,"JxPath":2447}
args = dotdict({
    'NlLen':NlLen_map[sys.argv[2]],
    'CodeLen':CodeLen_map[sys.argv[2]],
    'SentenceLen':10,
    'batch_size':60,
    'embedding_size':32,
    'WoLen':15,
    'Vocsize':100,
    'Nl_Vocsize':100,
    'max_step':3,
    'margin':0.5,
    'poolsize':50,
    'Code_Vocsize':100,
    'seed':0,
    'lr':1e-3
})
os.environ['PYTHONHASHSEED'] = str(args.seed)

def save_model(model, dirs = "checkpointcodeSearch"):
    if not os.path.exists(dirs):
        os.makedirs(dirs)
    torch.save(model.state_dict(), dirs + '/best_model.ckpt')


def load_model(model, dirs="checkpointcodeSearch"):
    assert os.path.exists(dirs + '/best_model.ckpt'), 'Weights for saved model not found'
    model.load_state_dict(torch.load(dirs + '/best_model.ckpt'))

use_cuda = torch.cuda.is_available()
#use_cuda=False
print("********************************",use_cuda)
def gVar(data):
    tensor = data
    if isinstance(data, np.ndarray):
        tensor = torch.from_numpy(data)
    elif isinstance(data, list):
        for i in range(len(data)):
            data[i] = gVar(data[i])
        tensor = data
    else:
        assert isinstance(tensor, torch.Tensor)
    if use_cuda:
        tensor = tensor.cuda()
    return tensor

def train(t = 5, p='Math'):

    torch.manual_seed(args.seed)
    np.random.seed(args.seed)  
    random.seed(args.seed + t)
    torch.cuda.manual_seed(args.seed)
    torch.cuda.manual_seed_all(args.seed) 

    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.deterministic = True
    dev_set = SumDataset(args, "test", p, testid=t)
    val_set = SumDataset(args, "val", p, testid=t)
    data = pickle.load(open(p + '.pkl', 'rb'))
    dev_data = pickle.load(open(p + '.pkl', 'rb'))
    train_set = SumDataset(args, "train", testid=t, proj=p, lst=dev_set.ids + val_set.ids)
    numt = len(train_set.data[0])
    args.Code_Vocsize = len(train_set.Code_Voc)
    args.Nl_Vocsize = len(train_set.Nl_Voc)
    args.Vocsize = len(train_set.Char_Voc)

    print(dev_set.ids)
    model = NlEncoder(args)
    if use_cuda:
        print('using GPU')
        model = model.cuda()
    maxl = 1e9
    optimizer = ScheduledOptim(optim.Adam(model.parameters(), lr=args.lr), args.embedding_size, 4000)
    maxAcc = 0
    minloss = 1e9
    rdic = {}
    brest = []
    bans = []
    batchn = []
    each_epoch_pred = {}
    for x in dev_set.Nl_Voc:
      rdic[dev_set.Nl_Voc[x]] = x
    testtime=datetime.now()-datetime.now()
    traintime=datetime.now()-datetime.now()
    for epoch in range(15):
        index = 0
        for dBatch in tqdm(train_set.Get_Train(args.batch_size)):
            if index == 0:
                TestStartTime = datetime.now()
                tmp={}
                accs = []
                loss = []
                model = model.eval()
                
                score2 = []
                for k, devBatch in tqdm(enumerate(val_set.Get_Train(len(val_set)))):
                        for i in range(len(devBatch)):
                            devBatch[i] = gVar(devBatch[i])
                        with torch.no_grad():
                            l, pre, _ = model(devBatch[0], devBatch[1], devBatch[2], devBatch[3], devBatch[4], devBatch[5], devBatch[6], devBatch[7])
                            resmask = torch.eq(devBatch[0], 2)
                            s = -pre#-pre[:, :, 1]
                            s = s.masked_fill(resmask == 0, 1e9)
                            pred = s.argsort(dim=-1)
                            pred = pred.data.cpu().numpy()
                            alst = []

                            for k in range(len(pred)): 
                                datat = data[val_set.ids[k]]
                                maxn = 1e9
                                lst = pred[k].tolist()[:resmask.sum(dim=-1)[k].item()]#score = np.sum(loss) / numt
                                tmp[val_set.ids[k]]=lst
                                #bans = lst
                                for x in datat['ans']:
                                    i = lst.index(x)
                                    maxn = min(maxn, i)
                                score2.append(maxn)

                TestEndTime = datetime.now()
                testtime+=TestEndTime-TestStartTime
                each_epoch_pred[epoch] = tmp
                score = score2[0]
                #print('curr accuracy is ' + str(score) + "," + str(score2))
                if score2[0] == 0:
                    batchn.append(epoch)
                    

                if  maxl >= score:
                    brest = score2
                    bans = tmp
                    maxl = score
                    #print("find better score " + str(score) + "," + str(score2))
                    #save_model(model)
                    #torch.save(model.state_dict(), os.path.join(wandb.run.dir, 'model.pt'))
                model = model.train()
            TrainStartTime = datetime.now()
            for i in range(len(dBatch)):
                dBatch[i] = gVar(dBatch[i])
            loss, _, _ = model(dBatch[0], dBatch[1], dBatch[2], dBatch[3], dBatch[4], dBatch[5], dBatch[6], dBatch[7])
            #print(loss.mean().item())
            optimizer.zero_grad()
            loss = loss.mean()
            loss.backward()

            optimizer.step_and_update_lr()
            index += 1
            TrainEndTime = datetime.now()
            traintime+=TrainEndTime-TrainStartTime
    # 打开Excel文
    workbook = openpyxl.load_workbook('time.xlsx')
    # 获取Sheet1工作表
    sheet = workbook['Sheet1']
    # 获取下一个空行的行数
    next_row = sheet.max_row + 1
    # 循环遍历每个单元格，并将其值设置为对应行的值
    sheet.cell(row=next_row, column=1).value = str(traintime)
    sheet.cell(row=next_row, column=2).value = str(testtime)
    sheet.cell(row=next_row, column=3).value = p
    # 保存Excel文件
    workbook.save('time.xlsx')
    return brest, bans, batchn, each_epoch_pred



if __name__ == "__main__":
    args.lr = float(sys.argv[3])
    args.seed = int(sys.argv[4])
    args.batch_size = int(sys.argv[5])
    np.set_printoptions(threshold=sys.maxsize)
    res = {}    
    p = sys.argv[2]
    res[int(sys.argv[1])] = train(int(sys.argv[1]), p)
    open('%sres%d_%d_%s_%s.pkl'%(p, int(sys.argv[1]), args.seed, args.lr, args.batch_size), 'wb').write(pickle.dumps(res))



